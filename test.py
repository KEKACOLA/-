import os
#import xlrd
import xlwt
import datetime
import openpyxl

value_ch1_1mon = []
value_ch1_15min = []
value_ch2_1mon = []
value_ch2_15min = []
value_ch3_1mon = []
value_ch3_15min = []
value_ch4_1mon = []
value_ch4_15min = []
value_ch5_1mon = []
value_ch5_15min = []
value_ch6_1mon = []
value_ch6_15min = []

def read_excel():
    start = datetime.datetime.now()
    print('仅支持.xlsx文件，请先确保格式正确')
    #打开文件
    path = input('input file path: ')

    #判断文件存在
    path = pathExists(path)

    workbook = openpyxl.load_workbook(str(path))
    
    #打开第一张表
    table = workbook.active

    #获取行列数
    nrows = table.max_row
    ncols = table.max_column
    #print(nrows, ncols)
    
    #第一行的时间
    firsttime = table.cell(3,1)
    firsttime = datetime.datetime.strptime(str(firsttime.value), "%Y-%m-%d %H:%M:%S")
    #print(firsttime)

    #最后一条数据所在行的时间
    endtime = table.cell(nrows,1)
    endtime = datetime.datetime.strptime(str(endtime.value), "%Y-%m-%d %H:%M:%S")
    print('endtime is : ' + str(endtime))

    #1个月前的日期
    starttime = endtime - datetime.timedelta(days=30)
    print('starttime should is: ' + str(starttime))

    #如果首行日期晚于计算出的起始日期，则从第一行开始统计
    if starttime < firsttime:
        print('first time actually is: ' + str(firsttime))
        startRow = 3
        getDataFor15min(startRow, table, endtime)
        
    #如果首行日期早于计算出的起始日期，则从计算日期所在行开始统计
    else:
        print('first time actually as same as starttime. is: ' + str(starttime))
        startRow = findRowByDate(starttime, nrows, table)
        print(startRow)
        #getDataFor15min(startRow,nrows-1,table)
        getDataFor15min(startRow, table, endtime)
        #print(startRow,nrows-1)
        
    big1 = max(value_ch1_1mon)
    big2 = max(value_ch2_1mon)
    big3 = max(value_ch3_1mon)
    big4 = max(value_ch4_1mon)
    big5 = max(value_ch5_1mon)
    big6 = max(value_ch6_1mon)
    print('CH01 ~ CH06 在近1个月的最大值分别是: '+ str(big1)+', '+str(big2)+', '+str(big3)+', '+str(big4)+', '+str(big5)+', '+str(big6))
    value_ch1_1mon.clear()
    value_ch2_1mon.clear()
    value_ch3_1mon.clear()
    value_ch4_1mon.clear()
    value_ch5_1mon.clear()
    value_ch6_1mon.clear()

    end = datetime.datetime.now()
    time = end-start
    print('program is running : '+ str(time))

    while True:
        content = input('输入【q + 回车】退出程序')
        if content == 'q':
            break
        

def getDataFor15min(startRow, table, endtime):    
    value = 0
    nrows = table.max_row
    print('row now is: '+ str(startRow))
    
    #所在行的时间
    row_start_time = table.cell(startRow,1)
    row_start_time = datetime.datetime.strptime(str(row_start_time.value), "%Y-%m-%d %H:%M:%S")
    #print(row_start_time)

    #所在行15分钟后的时间    
    aimTime = table.cell(startRow,1)
    aimTime = datetime.datetime.strptime(str(aimTime.value), "%Y-%m-%d %H:%M:%S")    
    datetime.timedelta(days=0,seconds=0,microseconds=0,milliseconds=0,minutes=0,hours=0,weeks=0)
    aimTime = aimTime + datetime.timedelta(minutes=15)
    #print(aimTime)


   
    #15分钟后所在行   
    if aimTime > endtime:
        row_after_15min = table.max_row
        print('aim later than endtime, row_after_15min is: '+str(row_after_15min)+'startRow is: '+str(startRow))
    else:
        row_after_15min = findRowByDate(aimTime,nrows,startRow,table)
    
    n_range = row_after_15min - startRow
    print('range is: ' + str(n_range))
    if(n_range <= 0):
        return

    #计算期间最大值
    global value_ch1_15min,value_ch2_15min,value_ch3_15min,value_ch4_15min,value_ch5_15min,value_ch6_15min
    global value_ch1_1mon,value_ch2_1mon,value_ch3_1mon,value_ch4_1mon,value_ch5_1mon,value_ch6_1mon
    for i in range(n_range):
        data = table.cell(startRow + i ,2)
        value_ch1_15min.append(data.value)
        
        data = table.cell(startRow + i ,3)
        value_ch2_15min.append(data.value)
        
        data = table.cell(startRow + i ,4)
        value_ch3_15min.append(data.value)
        
        data = table.cell(startRow + i ,5)
        value_ch4_15min.append(data.value)
        
        data = table.cell(startRow + i ,6)
        value_ch5_15min.append(data.value)
        
        data = table.cell(startRow + i ,7)
        value_ch6_15min.append(data.value)

    big_data1 = max(value_ch1_15min)
    big_data2 = max(value_ch2_15min)
    big_data3 = max(value_ch3_15min)
    big_data4 = max(value_ch4_15min)
    big_data5 = max(value_ch5_15min)
    big_data6 = max(value_ch6_15min)
    print('big data in 15min is: '+ str(big_data1)+','+ str(big_data2)+','+ str(big_data3)+','+ str(big_data4)+','+ str(big_data5)+','+ str(big_data6))
    
    
    #15min最大值放入数组
    value_ch1_1mon.append(big_data1)
    value_ch2_1mon.append(big_data2)
    value_ch3_1mon.append(big_data3)
    value_ch4_1mon.append(big_data4)
    value_ch5_1mon.append(big_data5)
    value_ch6_1mon.append(big_data6)
    
    value_ch1_15min.clear()
    value_ch2_15min.clear()
    value_ch3_15min.clear()
    value_ch4_15min.clear()
    value_ch5_15min.clear()
    value_ch6_15min.clear()
    
    #global value_1mon
    #value_1mon.append(value)
    #print(value_1mon)
    #value = 0
    
    return getDataFor15min(row_after_15min, table, endtime)

"""
    #计算期间平均值
    for i in range(n_range):
        #所在行第2列数值
        data =  table.cell(startRow + i ,2)
        value += data.value
    #print(value)

    #15min内的平均值
    value /= n_range
    print(value)
"""

#在首列查找字符串，定位获取首行数据
def findRowByDate(date,nrows,startRow,table):
    for x in range(startRow,nrows):
            if (x==0 or x==1 or x==2):
                continue
            #获得当前行的时间
            dt = table.cell(x,1)
            dt = datetime.datetime.strptime(str(dt.value), "%Y-%m-%d %H:%M:%S")
            if(dt == date):
                #获得计算开始行
                #print(x)
                break
    return x

def pathExists(path):
    re = os.path.exists(path)
    print(re)
    if re == True:
        print('loading...')
        return path
    else:
        path = input('input correct file path : ')
        pathExists(path)


#获得最后一行日期时间，确定一个月起止行
    #取15分钟内平均值
        #计算最大值
        
      
        

if __name__ == '__main__':
    read_excel()
